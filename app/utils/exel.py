import io
import os
import pandas as pd
import tempfile
from sqlalchemy.orm import Session
from app.database.models import User, AnonymousQuestion
from app.database.database import AsyncSession
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.repositories.catalog_repo import CatalogRepo
import re
from openpyxl.styles import Font, PatternFill

def user_create_excel_file(users: list[User]) -> bytes:
    # Словарь соответствия английских и русских названий полей
    field_mapping = {
        "Full Name": "ФИО",
        "Username": "Имя пользователя",
        "Telegram ID": "ID Telegram",
        "T-Points": "T-Points",
        "Department": "Отдел",
        "Post": "Должность",
        "Birth Date": "Дата рождения",
        "Hire Date": "Дата приема на работу",
        "Active": "Активен (1-да, 0-нет)"
    }
    
    # Словарь с описаниями полей и форматами
    field_descriptions = {
        "ФИО": "Полное имя сотрудника",
        "Имя пользователя": "Имя пользователя в Telegram (без символа @)",
        "ID Telegram": "Уникальный идентификатор Telegram пользователя (числовой)",
        "T-Points": "Количество T-Points на балансе пользователя",
        "Отдел": "Наименование отдела сотрудника",
        "Должность": "Должность сотрудника",
        "Дата рождения": "Дата рождения в формате ГГГГ-ММ-ДД",
        "Дата приема на работу": "Дата приема на работу в формате ГГГГ-ММ-ДД",
        "Активен (1-да, 0-нет)": "Флаг активности пользователя: 1 - активен, 0 - неактивен"
    }

    data = [
        {
            field_mapping["Full Name"]: user.fullname,
            field_mapping["Username"]: user.username or "",
            field_mapping["Telegram ID"]: user.telegram_id,
            field_mapping["T-Points"]: user.tpoints,
            field_mapping["Department"]: user.department or "",
            field_mapping["Post"]: user.post or "",
            field_mapping["Birth Date"]: user.birth_date,
            field_mapping["Hire Date"]: user.hire_date,
            field_mapping["Active"]: int(user.is_active),
        }
        for user in users
    ]
    
    df = pd.DataFrame(data)
    
    # Create description DataFrame
    desc_data = []
    for field, description in field_descriptions.items():
        desc_data.append({
            'Поле': field,
            'Описание': description
        })
    df_desc = pd.DataFrame(desc_data)
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Пользователи')
        df_desc.to_excel(writer, index=False, sheet_name='Инструкция')
        
        # Adjust column widths for users sheet
        worksheet = writer.sheets['Пользователи']
        for idx, col in enumerate(df.columns):
            column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(column_width, 50)
        
        # Adjust column widths for description sheet
        worksheet = writer.sheets['Инструкция']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 100
        
        # Style the headers in both sheets
        for sheet in [worksheet, writer.sheets['Пользователи']]:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    
    output.seek(0)
    return output.getvalue()


async def parse_excel_file(file_path: str, db: AsyncSession) -> dict:
    print(f"[DEBUG] parse_excel_file вызвана с путем: {file_path}")
    try:
        # Проверка существования файла
        if not os.path.exists(file_path):
            print(f"[ERROR] Файл не существует: {file_path}")
            return {"success": False, "message": "Файл не найден"}
            
        print(f"Начинаем обработку файла: {file_path}")
        
        # Словарь соответствия русских и английских названий полей
        reverse_field_mapping = {
            "ФИО": "Full Name",
            "Имя пользователя": "Username",
            "ID Telegram": "Telegram ID",
            "T-Points": "T-Points",
            "Отдел": "Department",
            "Должность": "Post",
            "Дата рождения": "Birth Date",
            "Дата приема на работу": "Hire Date",
            "Активен (1-да, 0-нет)": "Active"
        }
        
        # Читаем Excel с обработкой дат
        df = pd.read_excel(file_path, parse_dates=["Дата рождения", "Дата приема на работу", "Birth Date", "Hire Date"])
        print(f"Файл успешно прочитан. Найдено {len(df)} записей")
        print(f"Columns: {df.columns.tolist()}")
        
        # Выводим первые несколько строк для проверки
        print("Пример данных:")
        print(df.head(2))
        
        # Переименовываем столбцы с русского на английский для обработки
        rename_dict = {ru: en for ru, en in reverse_field_mapping.items() if ru in df.columns}
        if rename_dict:
            df = df.rename(columns=rename_dict)
        
        # Проверяем наличие всех необходимых столбцов
        required_columns = ["Full Name", "Telegram ID"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            # Переводим названия обратно на русский для понятной ошибки
            missing_ru = [list(reverse_field_mapping.keys())[list(reverse_field_mapping.values()).index(col)] for col in missing_columns]
            print(f"ОШИБКА: Отсутствуют столбцы: {', '.join(missing_ru)}")
            return {"success": False, "message": f"Отсутствуют столбцы: {', '.join(missing_ru)}"}
        
        # Счетчики
        updated = 0
        errors = 0
        error_messages = []
        
        print("Начинаем обработку записей...")
        # Обрабатываем каждую строку
        for index, row in df.iterrows():
            try:
                telegram_id = row["Telegram ID"]
                print(f"Обработка записи {index+1}/{len(df)}: Telegram ID {telegram_id}")
                
                # Используем select вместо query
                from sqlalchemy import select
                from app.database.models import User
                
                # Получаем пользователя по Telegram ID или создаем нового
                stmt = select(User).where(User.telegram_id == telegram_id)
                result = await db.execute(stmt)
                user = result.scalar_one_or_none()
                
                if not user:
                    print(f"Создаем нового пользователя с Telegram ID {telegram_id}")
                    user = User(telegram_id=telegram_id)
                    db.add(user)
                else:
                    print(f"Обновляем существующего пользователя с Telegram ID {telegram_id}")
                
                # Обновляем данные пользователя
                user.fullname = row["Full Name"]
                user.username = row["Username"] if not pd.isna(row.get("Username", None)) else None
                user.tpoints = int(row["T-Points"]) if not pd.isna(row.get("T-Points", None)) else 0
                user.department = row["Department"] if not pd.isna(row.get("Department", None)) else None
                user.post = row["Post"] if not pd.isna(row.get("Post", None)) else None
                user.birth_date = row["Birth Date"] if not pd.isna(row.get("Birth Date", None)) else None
                user.hire_date = row["Hire Date"] if not pd.isna(row.get("Hire Date", None)) else None
                user.is_active = bool(row["Active"]) if not pd.isna(row.get("Active", None)) else True
                
                print(f"Данные пользователя успешно обновлены")
                updated += 1
            except Exception as e:
                errors += 1
                error_message = f"Ошибка в строке {index+1} (Telegram ID {row.get('Telegram ID', 'неизвестно')}): {str(e)}"
                print(f"ОШИБКА: {error_message}")
                error_messages.append(error_message)
        
        # Сохраняем изменения в базе данных
        print(f"Сохраняем изменения в базе данных...")
        await db.commit()
        print(f"Изменения сохранены. Обновлено записей: {updated}, ошибок: {errors}")
        
        result = {
            "success": True,
            "updated": updated,
            "errors": errors
        }
        
        if error_messages:
            result["error_messages"] = error_messages
            
        print(f"Результат импорта: {result}")
        return result
    
    except Exception as e:
        print(f"КРИТИЧЕСКАЯ ОШИБКА при импорте: {str(e)}")
        await db.rollback()  # Асинхронный метод rollback
        return {"success": False, "message": f"Ошибка при импорте: {str(e)}"}
    
def anon_question_create_excel_file(questions: list[AnonymousQuestion]) -> bytes:
    data = []
    for question in questions:
        status_display = {
            None: "Новый",
            "read": "Прочитано",
            "later": "Отложено"
        }.get(question.question_status, question.question_status)

        data.append({
            "ID": question.id,
            "Вопрос": question.question_text,
            "Статус": status_display,
            "Дата создания": question.submitted_at.strftime('%Y-%m-%d %H:%M'),
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Анонимные вопросы')
        worksheet = writer.sheets['Анонимные вопросы']
        
        from openpyxl.utils import get_column_letter

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output.seek(0)
    return output.getvalue()
